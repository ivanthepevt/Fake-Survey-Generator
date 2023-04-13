#fake data geenrator
#author: Ivan
#contact: ivan.lolihunter@gmail.com

from openpyxl import Workbook
from openpyxl import load_workbook
from tkinter import *
from tkinter import filedialog as fd
import random
import math

def datagen(datanum, mean):
    if float(mean).is_integer():
        return [mean]*datanum
    else:
        base = int(math.floor(mean)*datanum)
        leak = int(mean*datanum - base)
        return (([math.floor(mean)]*(datanum-leak))+[math.floor(mean+1)]*leak)

def string_to_dict(s):
    return {k.strip(): v.strip() for k, v in [pair.split(':') for pair in s.split(' ')]}

def getnewdir(olddir):
    splitdir = olddir.split("/")
    oldfilename = splitdir[-1]
    splitdir[-1] = "processed_" + oldfilename
    return "/".join(splitdir)

def createlabel(string, n):
    result = ["Mean"]
    for i in range(1, n+1):
        result.append(f"{string}{i}")
    return result

filename = fd.askopenfilename()

print("File chosen: " + filename)

#indicate = input("Type your indicator format here (for example: Them 10 ul lan thu #) with # represent the timer: ")
input_string = "4*GV:-0.75 5*SV:0.75 3*GD:0.75 6*MT:-0.75 4*TL:0.75"
meanVars = string_to_dict(input_string.strip())
print(meanVars)
NamesOfMeans = list(meanVars.keys())

workbook = load_workbook(filename)

# Create a new workbook for the output
output_workbook = Workbook()
output_sheet = output_workbook.active

# Get the input sheet and find the last column with data
input_sheet = workbook.active
last_col = input_sheet.max_column

# Copy the first row of the input sheet to the output sheet and initialize the labels
first_row = input_sheet[1]
for cell in first_row:
    output_sheet[cell.coordinate] = cell.value
output_sheet.cell(row=1, column=last_col+1, value="Mean")
for i in range(len(NamesOfMeans)):
    output_sheet.cell(row=1, column=i+last_col+2, value=NamesOfMeans[i].split("*")[1])

# Iterate over the rows and calculate the average value for each row

for row in input_sheet.iter_rows(min_row=2, min_col=1, max_col=last_col, values_only=True):
    row_average = sum(row) / len(row)
    row = list(row) + [row_average]
    for akey in NamesOfMeans:
        meanName = akey.split("*")[1]
        gen_num = int(akey.split("*")[0])
        if float(meanVars[akey]) > 0:
            upperBound = math.ceil(gen_num * (row_average + float(meanVars[akey])))
            lowerBound = math.ceil(gen_num * row_average)
            thisMean = random.choice(range(lowerBound, upperBound + 1))/gen_num
        else:
            lowerBound = math.floor(gen_num * (row_average + float(meanVars[akey])))
            upperBound = math.floor(gen_num * row_average)
            thisMean = random.choice(range(lowerBound, upperBound + 1))/gen_num
        thisMean = max(0, min(thisMean, 5)) #make sure mean is larger than 0 and less than 5
        row = list(row) + [thisMean]
    output_sheet.append(row)

#generate more data from the means
new_last_col = output_sheet.max_column

for index, col in enumerate(output_sheet.iter_cols(min_col=last_col+2, max_col=new_last_col, min_row = 2, values_only=True), start=0):

    akey = NamesOfMeans[index]
    meanName = akey.split("*")[1]
    gen_num = int(akey.split("*")[0])
    new_sheet = output_workbook.create_sheet(meanName + "(s)")
    labels = createlabel(meanName, gen_num)
    new_sheet.append(labels)

    for aMean in col:
        fakeData = datagen(gen_num, aMean)
        random.shuffle(fakeData)
        new_sheet.append([aMean] + fakeData)

output_workbook.save(getnewdir(filename))